import logging
import subprocess
import traceback
import sys

import aiohttp
import asyncio
import pkg_resources

# Create a logger instance
logger = logging.getLogger("scraping")
logger.setLevel(logging.INFO)

# Create a console handler
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# Create a file handler
file_handler = logging.FileHandler("scraping.log")
file_handler.setLevel(logging.INFO)
file_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

# Upgrade pip
try:
    subprocess.check_call(['python', '-m', 'pip', 'install', '--upgrade', 'pip'])
except subprocess.CalledProcessError as e:
    logger.error("Error occurred while upgrading pip: {}".format(str(e)))
    exit(1)

# Check for missing libraries
missing_libraries = []
required_libraries = ['spacy', 'pytesseract', 'requests', 'bs4', 'tqdm', 'fake_useragent', 'opencv-python', 'python-Levenshtein', 'nltk', 'openpyxl', 'aiohttp', 'asyncio', 'pandas', 'openpyxl', 'fake_useragent']

for library in required_libraries:
    try:
        pkg_resources.get_distribution(library)
    except pkg_resources.DistributionNotFound:
        missing_libraries.append(library)

if missing_libraries:
    logger.error("The following libraries are missing or not installed:")
    for library in missing_libraries:
        logger.error(f"- {library}")

    response = input("\nDo you want to install the missing libraries? (y/n): ")
    if response.lower() == 'y':
        for library in missing_libraries:
            try:
                subprocess.check_call(['pip', 'install', library])
                logger.info(f"Library {library} has been installed.")
            except subprocess.CalledProcessError as e:
                logger.error(f"Error occurred while installing {library}.")
                logger.error(str(e))
    else:
        logger.error("Please install the missing libraries before running the script.")
        exit(1)
else:
    logger.info("All required libraries are installed.")

# Check for missing language modules
missing_modules = []
required_modules = ['en_core_web_sm']

for module in required_modules:
    try:
        pkg_resources.get_distribution(module)
    except pkg_resources.DistributionNotFound:
        missing_modules.append(module)

if missing_modules:
    logger.error("The following language modules are missing or not installed:")
    for module in missing_modules:
        logger.error(f"- {module}")

    response = input("\nDo you want to download and install the missing language modules? (y/n): ")
    if response.lower() == 'y':
        try:
            subprocess.check_call(['python', '-m', 'spacy', 'download'] + missing_modules)
            logger.info("Missing language modules have been downloaded and installed.")
        except subprocess.CalledProcessError:
            logger.error("Failed to download and install language modules.")
            logger.error("Please make sure 'spacy' is installed and try running the following command manually:")
            logger.error("python -m spacy download en_core_web_sm")
            exit(1)
    else:
        logger.error("Please install the missing language modules before running the script.")
        exit(1)
else:
    logger.info("All required language modules are installed.")

from spacy.util import get_lang_class
def prompt_additional_language_modules():
    # Hardcoded list of supported language models
    supported_languages = ['en', 'de', 'fr', 'es', 'it', 'nl']
    installed_languages = []
    for lang in supported_languages:
        try:
            get_lang_class(lang)
            installed_languages.append(lang)
        except OSError:
            pass

    if not installed_languages:
        logger.info("No additional language models are currently installed.")
        return

    logger.info("Installed language models:")
    for lang in installed_languages:
        logger.info(lang)

    response = input("Do you want to download and install additional language modules? (y/n): ")
    if response.lower() == 'y':
        language_code_pattern = r'^[a-z]{2}$'
        language_modules = []

        while True:
            language_code = input("Enter the language code (2 characters): ")
            if re.match(language_code_pattern, language_code):
                if language_code in installed_languages:
                    logger.info("Language model is already installed.")
                else:
                    language_modules.append(language_code)
                response = input("Do you want to add more language modules? (y/n): ")
                if response.lower() != 'y':
                    break
            else:
                logger.error("Invalid language code. Please enter a valid 2-character language code.")

        try:
            subprocess.check_call(['python', '-m', 'spacy', 'download'] + language_modules)
            logger.info("Additional language modules have been downloaded and installed.")
        except subprocess.CalledProcessError:
            logger.error("Failed to download and install additional language modules.")
            logger.error("Please make sure 'spacy' is installed and try running the following command manually:")
            logger.error("python -m spacy download {}".format(" ".join(language_modules)))
            sys.exit(1)
    else:
        logger.info("No additional language modules will be installed.")


# Prompt user for levenshtein_threshold
while True:
    try:
        logger.info("Levenshtein threshold: The Levenshtein threshold is used for measuring the difference between two strings. A lower threshold allows for more lenient matching. A suggested value is 3, but you can adjust it based on your requirements.")
        levenshtein_threshold = int(input("Enter the Levenshtein threshold for name matching (suggested value: 3): "))
        if levenshtein_threshold >= 0:
            break
        else:
            logger.error("Please enter a non-negative integer value.")
    except ValueError:
        logger.error("Invalid input. Please enter a non-negative integer value.")

# Prompt user for distance_threshold
while True:
    try:
        logger.info("Name threshold: The distance threshold is used for similarity matching between names and email addresses. It represents the minimum similarity required for a match. A higher threshold means stricter matching. A suggested value is 0.8, but you can adjust it based on your data and preferences.")
        name_threshold = float(input("Enter the distance threshold for similarity matching (suggested value: 0.8): "))
        if 0 <= name_threshold <= 1:
            break
        else:
            logger.error("Please enter a value between 0 and 1.")
    except ValueError:
        logger.error("Invalid input. Please enter a numeric value.")
import Levenshtein
import cv2
import numpy as np
from nltk import word_tokenize, pos_tag
import spacy

def calculate_distance(name, email, threshold):
    email_prefix = email.split("@")[0]
    distance = Levenshtein.distance(name.lower(), email_prefix.lower())
    return distance if distance <= threshold else float('inf')


def truncate_content(content, max_length):
    return content[:max_length] if len(content) > max_length else content
import fake_useragent
async def parse_subpages(content, base_url, target_netloc):
    try:
        processed_urls = set()
        user_agent = fake_useragent.UserAgent()

        async with aiohttp.ClientSession() as session:
            queue = asyncio.Queue()
            queue.put_nowait((base_url, content))

            while not queue.empty():
                current_url, current_content = await queue.get()

                if current_url in processed_urls:
                    continue

                processed_urls.add(current_url)

                headers = {'User-Agent': user_agent.random}
                async with session.get(current_url, headers=headers) as response:
                    status_code = response.status
                    content_type = response.headers.get('Content-Type', '')
                    content = await response.read()

                if status_code != 200:
                    logging.error(f"Failed to fetch URL: {current_url}")
                    continue

                if 'text/html' not in content_type:
                    logging.info("Skipping URL: {}".format(current_url))
                    continue

                max_content_length = 49141
                if len(content) > max_content_length:
                    logging.info("Truncating content for URL: {}".format(current_url))
                    content = content[:max_content_length]

                emails = extract_emails_from_content(content)
                logging.info("URL: {}".format(current_url))
                logging.info("Emails: {}".format(emails))

                names = [extract_names(content, email, name_threshold, levenshtein_threshold, truncated_content=content) for email in emails]
                logging.info("Names: {}".format(names))

                yield current_url

                subpages = parse_subpages(content, current_url, target_netloc)
                async for subpage_url in subpages:
                    yield subpage_url

        return

    except aiohttp.ClientError as e:
        logging.error("Error occurred while accessing URL: {}\nError details: {}".format(base_url, str(e)))
        return

    except Exception as e:
        logging.exception(f"Unexpected error occurred while parsing subpages: {str(e)}")
        return


import re


email_patterns = [
    # Standard email format
    r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',

    # Obfuscated email patterns
    r'\b[A-Za-z0-9._%+-]+\[at\][A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
    r'\b[A-Za-z0-9._%+-]+\s*\*\s*[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
    r'\b[A-Za-z0-9]+\s*\[\.\]\s*[A-Za-z0-9]+\s*\[\.\]\s*[A-Za-z0-9]+\b',
    r'\b[A-Za-z0-9]+\s*\(\s*[A-Za-z0-9]+\s*\)\s*[A-Za-z0-9]+\b',
    r'\b[A-Za-z0-9]+\s*(\*|at|\[at\]|\(at\)|{at}|\[.\])\s*[A-Za-z0-9]+\b',

    # ASCII representation
    r'&#[0-9]+;',

    # JavaScript encoded emails
    r'unescape\(("[^"]+"|\'[^\']+\')\)',
    r'String\.fromCharCode\(([0-9]+,?)+\)',

    # Image-based obfuscation
    r'<img[^>]+src="data:image\/[^;]+;base64,[^"]+"[^>]*>',
    r'<img[^>]+src=\'data:image\/[^;]+;base64,[^\']+\'[^>]*>',

    # Hexadecimal representation
    r'[0-9a-fA-F]+',

    # Rot13 encoding
    r'[a-zA-Z](?:[^a-zA-Z]*[a-zA-Z]){3,}',
]

email_pattern = re.compile('|'.join(email_patterns), re.IGNORECASE)

def extract_emails_from_content(content):
    if isinstance(content, bytes):
        content = content.decode('utf-8', errors='ignore')

    soup = BeautifulSoup(content, 'html.parser')

    emails = []
    for pattern in email_patterns:
        try:
            extracted_emails = re.findall(pattern, content)
            for email in extracted_emails:
                if isinstance(email, str) and is_valid_email(email):
                    emails.append(email)
        except Exception as e:
            logging.error(f"Error occurred while extracting emails with pattern {pattern}: {str(e)}")

    email_tags = soup.select('a[href^="mailto:"]')
    for email_tag in email_tags:
        try:
            email = email_tag.get('href').replace('mailto:', '')
            if isinstance(email, str) and is_valid_email(email):
                emails.append(email.replace('[at]', '@'))
        except Exception as e:
            logging.error(f"Error occurred while extracting email from mailto tag: {str(e)}")

    return emails




def load_language_module(module_name):
    try:
        nlp = spacy.load(module_name)
        return nlp
    except OSError as e:
        logging.error(f"Failed to load language module '{module_name}': {str(e)}")
        sys.exit(1)
nlp = load_language_module('en_core_web_sm')


import re
import pytesseract


def extract_names(content, email, name_threshold, levenshtein_threshold, truncated_content=None, nlp=None):
    try:
        email_prefix = email.split("@")[0]

        if truncated_content:
            if isinstance(truncated_content, bytes):
                truncated_content = truncated_content.decode('utf-8', errors='ignore')
            doc = nlp(truncated_content)
        else:
            doc = nlp(content, disable=["parser"])

        names = []
        for entity in doc.ents:
            if entity.label_ == 'PERSON':
                name = entity.text.strip()
                if len(name) > name_threshold:
                    names.append(name)

        if names:
            best_match = min(names, key=lambda name: calculate_distance(name, email_prefix, levenshtein_threshold))
            return best_match

        words = []
        pos_tags = []
        for sentence in doc.sents:
            sentence_words = word_tokenize(sentence.text)
            sentence_pos_tags = pos_tag(sentence_words)
            words.extend(sentence_words)
            pos_tags.extend(sentence_pos_tags)

        names_nnp = [word for word, pos in pos_tags if pos == 'NNP' and len(word) > name_threshold]
        if names_nnp:
            best_match = min(names_nnp, key=lambda name: calculate_distance(name, email_prefix, levenshtein_threshold))
            return best_match

        return 'N/A'
    except Exception as e:
        logging.error("Error occurred while extracting names: {}".format(str(e)))
        return 'N/A'



def is_valid_email(email):
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    try:
        if isinstance(email, str):
            return bool(re.fullmatch(email_pattern, email))
        else:
            raise ValueError("Invalid email format. Email must be a string.")
    except Exception as e:
        logging.error(f"Error occurred while validating email: {email}\nError details: {str(e)}")
        return False



def extract_email_from_image(image_data):
    try:
        nparr = np.frombuffer(image_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)

        ret, thresh = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

        extracted_text = pytesseract.image_to_string(thresh, config='--psm 6')

        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        extracted_emails = re.findall(email_pattern, extracted_text)

        if extracted_emails:
            return extracted_emails[0]
        else:
            return ''
    except Exception as e:
        logging.error("Error occurred while extracting email from image: {}".format(str(e)))
        return ''
import logging
from urllib.parse import urlparse, urljoin
from bs4 import BeautifulSoup

async def fetch(session, url, user_agent):
    headers = {'User-Agent': user_agent}
    try:
        async with session.get(url, headers=headers) as response:
            content_type = response.headers.get('Content-Type', '')
            content = await response.read()
            return response.status, content_type, content

    except aiohttp.ClientError as e:
        logging.error(f"Error occurred while fetching URL: {url}\nError details: {str(e)}")
        raise

    except Exception as e:
        logging.exception(f"Unexpected error occurred while fetching URL: {url}")
        raise


async def scrape_url(session, url, name_threshold, user_agent, target_netlocs, retries=3):
    for attempt in range(retries):
        try:
            emails, names, subpages, base_url, content = await scrape_data(url, name_threshold, user_agent, target_netlocs)
            return emails, names, subpages, base_url, content

        except aiohttp.ClientError as e:
            logging.error(f"Error occurred while accessing URL: {url}\nError details: {str(e)}")

        except Exception as e:
            logging.exception(f"Unexpected error occurred while scraping URL: {url}")
            raise

        if attempt < retries - 1:
            logging.info(f"Retrying {url} (attempt {attempt + 2} of {retries})...")
            await asyncio.sleep(1)  # Add a small delay before retrying

    logging.error(f"Failed to scrape {url} after {retries} attempts.")
    return [], [], set(), '', ''

async def scrape_data(urls_to_scrape, name_threshold, user_agent, target_netlocs, max_threads):
    all_emails = []
    email_to_name = {}
    processed_urls = set(urls_to_scrape)

    async def process_url(session, url):
        try:
            emails, names, subpages, base_url, content = await scrape_url(session, url, name_threshold, user_agent, target_netlocs)
            all_emails.extend([(email, url) for email in emails])
            for email, name in zip(emails, names):
                if is_valid_email(email):
                    email_to_name[email] = name

            return subpages

        except Exception as e:
            logging.exception(f"Error occurred during data scraping for URL: {url}")
            return []

    async def process_urls(urls_to_process):
        async with aiohttp.ClientSession() as session:
            tasks = [process_url(session, url) for url in urls_to_process]
            subpages_list = await asyncio.gather(*tasks, return_exceptions=True)
            subpages = set()

            for sublist in subpages_list:
                if isinstance(sublist, list):
                    subpages.update(sublist)

            return subpages

    try:
        while urls_to_scrape:
            subpages = await process_urls(urls_to_scrape)
            urls_to_scrape = subpages.difference(processed_urls)
            processed_urls.update(urls_to_scrape)

        if not all_emails:
            logging.warning("No valid emails found.")
        else:
            logging.info(f"Found {len(all_emails)} valid emails.")

    except Exception as e:
        logging.exception("Error occurred in the scrape_data function")

    return all_emails, email_to_name





import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook, Workbook
def save_data(all_emails, email_to_name):
    unique_emails = set()
    filtered_emails = []

    for email, url in all_emails:
        if email not in unique_emails:
            unique_emails.add(email)
            filtered_emails.append((email, email_to_name.get(email, 'N/A'), url))

    valid_emails = [(email, name, url) for email, name, url in filtered_emails if is_valid_email(email)]

    if not valid_emails:
        logging.warning("No valid emails found.")

    df = pd.DataFrame(valid_emails, columns=['Email', 'Name', 'URL'])

    try:
        # Load the workbook or create a new one
        try:
            workbook = load_workbook('scraped_data.xlsx')
        except FileNotFoundError:
            workbook = Workbook()

        # Select the active sheet or create a new one
        sheet_name = 'Sheet1'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Determine the starting row for writing data
        start_row = sheet.max_row + 1 if sheet.max_row > 0 else 1

        # Write data to the sheet
        for _, row in df.iterrows():
            sheet.append(row.tolist())

        # Move the "Name" column before the "URL" column
        if len(valid_emails) > 0:
            sheet.move_range(f'C{start_row}:C{sheet.max_row}', rows=-1)

        # Apply formatting
        headers = ['Email', 'Name', 'URL']
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Save the workbook
        workbook.save('scraped_data.xlsx')

    except Exception as e:
        logging.exception("Error occurred while saving data to Excel file:\n{}".format(str(e)))

    logging.info("Data saved to scraped_data.xlsx")



def main():
    try:
        # Prompt user for modifying user agent behavior
        user_agent_behavior = input("Do you want to modify the user agent behavior? (yes/no): ")

        # Prompt user for the URL to scrape
        urls = input("Enter the URLs to scrape (separated by commas if multiple): ").split(',')

        # Prompt user for the maximum depth of scraping
        while True:
            try:
                max_depth = int(input("Enter the maximum depth of scraping (recommended: 1-5): "))
                if max_depth > 0:
                    break
                else:
                    print("Please enter a positive integer value.")
            except ValueError:
                print("Invalid input. Please enter a positive integer value.")

        # Prompt user for the number of concurrent threads
        while True:
            try:
                max_threads = int(input("Enter the maximum number of concurrent threads (recommended: 5-10): "))
                if max_threads > 0:
                    break
                else:
                    print("Please enter a positive integer value.")
            except ValueError:
                print("Invalid input. Please enter a positive integer value.")

        # Prompt user for additional language modules
        prompt_additional_language_modules()

        # Get the netlocs of the URLs
        target_netlocs = [urlparse(url).netloc for url in urls]

        # Create an event loop
        loop = asyncio.get_event_loop()

        # Scrape data
        all_emails, email_to_name = loop.run_until_complete(scrape_data(urls, name_threshold, user_agent_behavior, target_netlocs, max_threads))

        # Save data
        save_data(all_emails, email_to_name)

    except Exception as e:
        logging.exception("An error occurred during execution: {}".format(str(e)))

if __name__ == '__main__':
    main()
